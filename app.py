"""
Sistema de Propostas Comerciais - Azevedo Contabilidade
Rode com: streamlit run app.py
"""

import streamlit as st
import json
import os
import requests
import io
from datetime import datetime, date
from gerar_proposta import gerar_docx
from sheets_db import (
    sheets_disponivel, load_propostas, save_proposta,
    update_proposta_status, delete_proposta,
    load_config_sheets, save_config_sheets,
    update_servicos_detalhados, expirar_itens_pendentes,
    update_proposta_autentique
)

# ===== CONFIG =====
st.set_page_config(
    page_title="Propostas - Azevedo Contabilidade",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed"
)

LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")
USING_SHEETS = sheets_disponivel()

# ===== AUTENTIQUE API =====
AUTENTIQUE_API_URL = "https://api.autentique.com.br/v2/graphql"
AUTENTIQUE_TOKEN = os.environ.get(
    "AUTENTIQUE_TOKEN",
    "25746e6bd578d8b5d11713f69027e686c5e07c738d2b9a9b71c1df419f37a4ad"
)
AUTENTIQUE_EMAIL_SIGNATARIO = "allan@azevedocontabilidade.com.br"

def enviar_para_autentique(docx_bytes, nome_documento, nome_cliente, email_signatario=AUTENTIQUE_EMAIL_SIGNATARIO):
    """
    Envia documento para Autentique para assinatura.
    - Signatário 1: Allan (via email) — SIGN
    - Signatário 2: Cliente (via link, só nome) — APPROVE
    Retorna (sucesso, dados) onde dados contém id do documento e link do cliente.
    """
    mutation = """
    mutation CreateDocumentMutation(
        $document: DocumentInput!,
        $signers: [SignerInput!]!,
        $file: Upload!
    ) {
        createDocument(
            document: $document,
            signers: $signers,
            file: $file
        ) {
            id
            name
            signatures {
                public_id
                name
                email
                action { name }
                link { short_link }
            }
        }
    }
    """

    variables = {
        "document": {
            "name": nome_documento,
            "message": f"Proposta comercial para {nome_cliente} - Azevedo Contabilidade",
            "reminder": "WEEKLY",
            "footer": "BOTTOM",
            "refusable": True,
            "sortable": True,
            "locale": {
                "country": "BR",
                "language": "pt-BR",
                "timezone": "America/Sao_Paulo",
                "date_format": "DD_MM_YYYY"
            }
        },
        "signers": [
            {
                "email": email_signatario,
                "action": "SIGN"
            },
            {
                "name": nome_cliente,
                "action": "APPROVE"
            }
        ],
        "file": None
    }

    operations = json.dumps({
        "query": mutation,
        "variables": variables
    })

    files_map = json.dumps({"file": ["variables.file"]})

    try:
        resp = requests.post(
            AUTENTIQUE_API_URL,
            headers={"Authorization": f"Bearer {AUTENTIQUE_TOKEN}"},
            data={
                "operations": operations,
                "map": files_map
            },
            files={
                "file": (f"{nome_documento}.docx", io.BytesIO(docx_bytes), "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
            },
            timeout=30
        )

        if resp.status_code != 200:
            return False, f"Erro HTTP {resp.status_code}: {resp.text[:200]}"

        data = resp.json()

        if "errors" in data:
            return False, f"Erro API: {data['errors'][0].get('message', str(data['errors']))}"

        doc_data = data.get("data", {}).get("createDocument", {})
        doc_id = doc_data.get("id", "")

        # Encontrar o link do cliente (signatário sem email = via link)
        def _extrair_link(sigs):
            for sig in sigs:
                action_name = (sig.get("action") or {}).get("name", "")
                if action_name == "APPROVE" or (not sig.get("email")):
                    link_info = sig.get("link") or {}
                    sl = link_info.get("short_link", "")
                    if sl:
                        return sl
            return ""

        link_cliente = _extrair_link(doc_data.get("signatures", []))

        # Autentique nem sempre retorna o short_link na mutation createDocument
        # (gerado de forma assíncrona). Se vazio, faz query subsequente até 3x.
        if not link_cliente and doc_id:
            for _tentativa in range(3):
                try:
                    import time as _t
                    _t.sleep(1.2)
                    q = 'query { document(id: "%s") { id signatures { name email action { name } link { short_link } } } }' % doc_id
                    r2 = requests.post(
                        AUTENTIQUE_API_URL,
                        headers={
                            "Authorization": f"Bearer {AUTENTIQUE_TOKEN}",
                            "Content-Type": "application/json"
                        },
                        json={"query": q},
                        timeout=15
                    )
                    if r2.status_code == 200:
                        d2 = r2.json().get("data", {}).get("document", {}) or {}
                        link_cliente = _extrair_link(d2.get("signatures", []))
                        if link_cliente:
                            break
                except Exception:
                    pass

        return True, {
            "autentique_id": doc_id,
            "autentique_link": link_cliente,
            "signatures": doc_data.get("signatures", [])
        }

    except requests.exceptions.Timeout:
        return False, "Timeout: o Autentique demorou para responder."
    except Exception as e:
        return False, f"Erro ao enviar: {str(e)}"


def consultar_autentique(doc_id):
    """Consulta status do documento no Autentique"""
    query = """
    query {
        document(id: "%s") {
            id
            name
            signatures {
                public_id
                name
                email
                action { name }
                link { short_link }
                signed { created_at }
                viewed { created_at }
                rejected { created_at }
            }
            files { original signed }
        }
    }
    """ % doc_id

    try:
        resp = requests.post(
            AUTENTIQUE_API_URL,
            headers={
                "Authorization": f"Bearer {AUTENTIQUE_TOKEN}",
                "Content-Type": "application/json"
            },
            json={"query": query},
            timeout=15
        )
        if resp.status_code == 200:
            data = resp.json()
            return True, data.get("data", {}).get("document", {})
        return False, f"Erro HTTP {resp.status_code}"
    except Exception as e:
        return False, str(e)

# ===== DATABASE (fallback local + Google Sheets) =====
DB_FILE = os.path.join(os.path.dirname(__file__), "propostas_db.json")
CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")

def load_db():
    if USING_SHEETS:
        return load_propostas()
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_db(data):
    if not USING_SHEETS:
        with open(DB_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

def load_config():
    if USING_SHEETS:
        return load_config_sheets()
    default = {"meta_mensal": 12000, "vendedores": ["Allan"]}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            for k, v in default.items():
                if k not in cfg:
                    cfg[k] = v
            return cfg
    save_config(default)
    return default

def save_config(cfg):
    if USING_SHEETS:
        save_config_sheets(cfg)
    else:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)

def fc(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def get_avatar_html(nome, fotos_dict=None):
    """Gera HTML de avatar: foto se disponível, ou iniciais coloridas."""
    if fotos_dict and nome in fotos_dict and fotos_dict[nome]:
        return f'<img src="{fotos_dict[nome]}" class="ranking-avatar" alt="{nome}">'
    # Gerar cor baseada no nome
    cores = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899', '#06b6d4', '#f97316']
    cor = cores[sum(ord(c) for c in nome) % len(cores)]
    iniciais = "".join(p[0].upper() for p in nome.split()[:2]) if nome else "?"
    return f'<div class="ranking-initials" style="background:{cor};">{iniciais}</div>'

# ===== CUSTOM CSS =====
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    .main .block-container { padding-top: 1rem; max-width: 1200px; }

    .stTabs [data-baseweb="tab-list"] {
        gap: 0;
        background: #f8f9fa;
        border-radius: 12px;
        padding: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 10px 24px;
        font-weight: 600;
        border-radius: 8px;
        font-size: 14px;
    }
    .stTabs [aria-selected="true"] {
        background: white !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }

    .header-bar {
        background: linear-gradient(135deg, #1a2744 0%, #2a3a5a 100%);
        border-radius: 16px;
        padding: 24px 32px;
        margin-bottom: 24px;
        display: flex;
        align-items: center;
        gap: 20px;
    }
    .header-bar h1 {
        color: white;
        font-size: 22px;
        font-weight: 700;
        margin: 0;
        font-family: 'Inter', sans-serif;
    }
    .header-bar p {
        color: #b8960c;
        font-size: 13px;
        margin: 4px 0 0;
        font-weight: 500;
    }

    .metric-card {
        background: white;
        border-radius: 16px;
        padding: 24px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
        border: 1px solid #f0f0f0;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }
    .metric-value {
        font-size: 28px;
        font-weight: 800;
        color: #1a2744;
        font-family: 'Inter', sans-serif;
        line-height: 1.2;
    }
    .metric-label {
        font-size: 11px;
        color: #9ca3af;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 600;
        margin-bottom: 8px;
    }
    .metric-icon {
        font-size: 20px;
        margin-bottom: 8px;
    }

    .meta-card {
        background: linear-gradient(135deg, #faf6e6 0%, #fff8e1 100%);
        border-radius: 16px;
        padding: 24px;
        border: 2px solid #b8960c33;
    }
    .meta-title {
        font-size: 14px;
        font-weight: 700;
        color: #b8960c;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 16px;
    }
    .meta-valor {
        font-size: 36px;
        font-weight: 800;
        color: #1a2744;
        font-family: 'Inter', sans-serif;
    }
    .meta-sub {
        font-size: 13px;
        color: #6b7280;
    }

    .progress-container {
        background: #e5e7eb;
        border-radius: 100px;
        height: 14px;
        overflow: hidden;
        margin: 12px 0;
    }
    .progress-bar {
        height: 100%;
        border-radius: 100px;
        transition: width 0.5s ease;
    }
    .progress-green { background: linear-gradient(90deg, #10b981, #34d399); }
    .progress-gold { background: linear-gradient(90deg, #b8960c, #d4af37); }
    .progress-red { background: linear-gradient(90deg, #ef4444, #f87171); }

    .ranking-row {
        display: flex;
        align-items: center;
        padding: 12px 16px;
        background: white;
        border-radius: 10px;
        margin-bottom: 8px;
        border: 1px solid #f0f0f0;
    }
    .ranking-pos {
        width: 32px;
        height: 32px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 800;
        font-size: 14px;
        margin-right: 14px;
    }
    .ranking-1 { background: #fef3c7; color: #b8960c; }
    .ranking-2 { background: #e5e7eb; color: #6b7280; }
    .ranking-3 { background: #fde8d0; color: #c2703e; }
    .ranking-other { background: #f3f4f6; color: #9ca3af; }
    .ranking-avatar {
        width: 36px;
        height: 36px;
        border-radius: 50%;
        margin-right: 12px;
        object-fit: cover;
        border: 2px solid #e5e7eb;
    }
    .ranking-initials {
        width: 36px;
        height: 36px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 700;
        font-size: 14px;
        color: white;
        margin-right: 12px;
        flex-shrink: 0;
    }
    .ranking-name {
        flex: 1;
        font-weight: 600;
        color: #1a2744;
        font-size: 14px;
    }
    .ranking-stats {
        text-align: right;
        font-size: 12px;
        color: #6b7280;
    }
    .ranking-valor {
        font-weight: 700;
        color: #1a2744;
        font-size: 15px;
    }

    .status-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 100px;
        font-size: 12px;
        font-weight: 600;
    }
    .status-enviada { background: #dbeafe; color: #2563eb; }
    .status-fechou { background: #d1fae5; color: #059669; }
    .status-fechou-parcial { background: #e0f2fe; color: #0284c7; }
    .status-nao-fechou { background: #fee2e2; color: #dc2626; }
    .status-pendente { background: #fef3c7; color: #d97706; }
    .status-expirado { background: #f3f4f6; color: #6b7280; }

    .svc-item {
        display: flex;
        align-items: center;
        padding: 10px 14px;
        background: #f9fafb;
        border-radius: 8px;
        margin-bottom: 6px;
        border: 1px solid #e5e7eb;
        gap: 12px;
    }
    .svc-item-desc { flex: 2; font-size: 13px; font-weight: 500; color: #374151; }
    .svc-item-valor { flex: 1; font-size: 13px; font-weight: 700; color: #1a2744; text-align: right; }
    .svc-item-per { flex: 1; font-size: 12px; color: #6b7280; text-align: center; }
    .svc-status-aprovado { color: #059669; font-weight: 700; }
    .svc-status-recusado { color: #dc2626; font-weight: 700; }
    .svc-status-pendente { color: #d97706; font-weight: 700; }
    .svc-status-expirado { color: #6b7280; font-weight: 700; font-style: italic; }

    .section-title {
        font-size: 13px;
        font-weight: 700;
        color: #b8960c;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        margin: 24px 0 16px;
        padding-bottom: 8px;
        border-bottom: 2px solid #f0f0f0;
    }

    /* Highlight / Spotlight Cards */
    .highlight-card {
        background: white;
        border-radius: 16px;
        padding: 20px 24px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
        border: 1px solid #f0f0f0;
        position: relative;
        overflow: hidden;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    .highlight-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 24px rgba(0,0,0,0.1);
    }
    .highlight-card::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 4px;
    }
    .hl-gold::before { background: linear-gradient(90deg, #b8960c, #d4af37); }
    .hl-green::before { background: linear-gradient(90deg, #10b981, #34d399); }
    .hl-blue::before { background: linear-gradient(90deg, #3b82f6, #60a5fa); }
    .hl-purple::before { background: linear-gradient(90deg, #8b5cf6, #a78bfa); }
    .hl-icon {
        width: 44px; height: 44px;
        border-radius: 12px;
        display: flex; align-items: center; justify-content: center;
        font-size: 22px;
        margin-bottom: 12px;
    }
    .hl-icon-gold { background: #fef3c7; }
    .hl-icon-green { background: #d1fae5; }
    .hl-icon-blue { background: #dbeafe; }
    .hl-icon-purple { background: #ede9fe; }
    .hl-label {
        font-size: 11px;
        color: #9ca3af;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-weight: 600;
        margin-bottom: 4px;
    }
    .hl-value {
        font-size: 20px;
        font-weight: 800;
        color: #1a2744;
        font-family: 'Inter', sans-serif;
        line-height: 1.2;
        margin-bottom: 2px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .hl-sub {
        font-size: 12px;
        color: #6b7280;
        line-height: 1.4;
    }
    .hl-sub strong { color: #1a2744; }
    .delta-up { color: #10b981; font-weight: 700; font-size: 12px; }
    .delta-down { color: #ef4444; font-weight: 700; font-size: 12px; }
    .delta-neutral { color: #9ca3af; font-weight: 600; font-size: 12px; }

    /* Insight box */
    .insight-box {
        background: linear-gradient(135deg, #f0f4ff 0%, #e8f0fe 100%);
        border-radius: 12px;
        padding: 16px 20px;
        border-left: 4px solid #3b82f6;
        margin-top: 8px;
    }
    .insight-box p {
        margin: 0;
        font-size: 13px;
        color: #374151;
        line-height: 1.5;
    }
    .insight-box strong { color: #1a2744; }

    /* Funnel */
    .funnel-step {
        display: flex;
        align-items: center;
        gap: 12px;
        margin-bottom: 10px;
    }
    .funnel-bar-bg {
        flex: 1;
        background: #f3f4f6;
        border-radius: 8px;
        height: 36px;
        overflow: hidden;
        position: relative;
    }
    .funnel-bar {
        height: 100%;
        border-radius: 8px;
        display: flex;
        align-items: center;
        padding-left: 12px;
        font-weight: 700;
        font-size: 13px;
        color: white;
        transition: width 0.6s ease;
    }
    .funnel-label {
        width: 100px;
        font-size: 12px;
        font-weight: 600;
        color: #374151;
        text-align: right;
    }
    .funnel-count {
        width: 40px;
        font-size: 13px;
        font-weight: 700;
        color: #1a2744;
        text-align: right;
    }

    div[data-testid="stForm"] {
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        padding: 24px;
        background: white;
    }
</style>
""", unsafe_allow_html=True)

# ===== HEADER =====
st.markdown("""
<div class="header-bar">
    <div>
        <h1>📄 Sistema de Propostas Comerciais</h1>
        <p>Azevedo Contabilidade — Contabilidade Estratégica & Planejamento Tributário</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ===== TABS =====
tab_dash, tab_nova, tab_hist, tab_comissao, tab_config = st.tabs(["📊 Dashboard", "📄 Nova Proposta", "📋 Histórico", "💰 Comissões", "⚙️ Configurações"])

# ==========================================
# TAB: DASHBOARD
# ==========================================
with tab_dash:
    db = load_db()
    config = load_config()
    meta_mensal = config.get("meta_mensal", 50000)

    # ===== HELPER: extrair itens aprovados com data_aprovacao =====
    def _parse_data(data_str):
        """Converte data para formato YYYY-MM-DD se necessário"""
        if not data_str:
            return ""
        data_str = str(data_str).strip()
        # Já no formato correto
        if len(data_str) >= 10 and data_str[4] == "-":
            return data_str[:10]
        # Formato DD/MM/YYYY
        if "/" in data_str:
            parts = data_str.split("/")
            if len(parts) == 3 and len(parts[2]) == 4:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        return data_str

    def extrair_itens_aprovados(propostas):
        """Retorna lista de dicts com: vendedor, valor, data_aprovacao, descricao"""
        itens = []
        for p in propostas:
            sd = p.get("servicos_detalhados", "")
            vendedor = p.get("vendedor", "Sem vendedor") or "Sem vendedor"
            data_proposta = _parse_data(p.get("data", ""))

            if sd and sd.strip().startswith("["):
                try:
                    svcs = json.loads(sd)
                    itens_aprovados = [s for s in svcs if s.get("status") == "Aprovado"]
                    soma_itens = sum(s.get("valor", 0) for s in itens_aprovados)
                    valor_proposta = p.get("valor", 0)

                    if itens_aprovados and soma_itens > 0:
                        # Itens têm valores individuais — usar normalmente
                        for s in itens_aprovados:
                            itens.append({
                                "vendedor": vendedor,
                                "valor": s.get("valor", 0),
                                "data_aprovacao": _parse_data(s.get("data_aprovacao", "")) or data_proposta,
                                "descricao": s.get("descricao", ""),
                                "cliente": p.get("cliente", "-"),
                                "periodicidade": s.get("periodicidade", "-")
                            })
                    elif itens_aprovados and soma_itens == 0 and valor_proposta > 0:
                        # Itens aprovados com valor 0 mas proposta tem valor — usar valor da proposta
                        itens.append({
                            "vendedor": vendedor,
                            "valor": valor_proposta,
                            "data_aprovacao": _parse_data(itens_aprovados[0].get("data_aprovacao", "")) or data_proposta,
                            "descricao": p.get("servicos", "-"),
                            "cliente": p.get("cliente", "-"),
                            "periodicidade": "-"
                        })
                    elif itens_aprovados:
                        # Itens aprovados mas tudo zerado — registra pelo menos o item
                        for s in itens_aprovados:
                            itens.append({
                                "vendedor": vendedor,
                                "valor": 0,
                                "data_aprovacao": _parse_data(s.get("data_aprovacao", "")) or data_proposta,
                                "descricao": s.get("descricao", ""),
                                "cliente": p.get("cliente", "-"),
                                "periodicidade": s.get("periodicidade", "-")
                            })
                except Exception:
                    # Se o JSON der erro, trata como proposta antiga
                    if p.get("status") in ("Fechou", "Fechou Parcial") and p.get("valor", 0) > 0:
                        itens.append({
                            "vendedor": vendedor,
                            "valor": p.get("valor", 0),
                            "data_aprovacao": data_proposta,
                            "descricao": p.get("servicos", "-"),
                            "cliente": p.get("cliente", "-"),
                            "periodicidade": "-"
                        })
            else:
                # Proposta antiga sem servicos_detalhados — valor integral
                if p.get("status") in ("Fechou", "Fechou Parcial") and p.get("valor", 0) > 0:
                    itens.append({
                        "vendedor": vendedor,
                        "valor": p.get("valor", 0),
                        "data_aprovacao": data_proposta,
                        "descricao": p.get("servicos", "-"),
                        "cliente": p.get("cliente", "-"),
                        "periodicidade": "-"
                    })
        return itens

    todos_itens_aprov = extrair_itens_aprovados(db)

    # ===== SELETOR DE MÊS =====
    st.markdown('<div class="section-title">📅 Período</div>', unsafe_allow_html=True)
    col_sel_mes, col_sel_ano, _ = st.columns([1, 1, 3])
    meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    with col_sel_mes:
        mes_sel = st.selectbox("Mês", range(1, 13), index=date.today().month - 1,
                               format_func=lambda x: meses_nomes[x-1], key="dash_mes")
    with col_sel_ano:
        ano_sel = st.number_input("Ano", min_value=2024, max_value=2030, value=date.today().year, key="dash_ano")

    prefixo_mes = f"{ano_sel}-{mes_sel:02d}"

    # Filtrar itens aprovados pelo mês da APROVAÇÃO (não da proposta)
    itens_mes = [i for i in todos_itens_aprov if i["data_aprovacao"].startswith(prefixo_mes)]
    receita_mes = sum(i["valor"] for i in itens_mes)

    # Propostas criadas no mês selecionado (para contagem)
    db_mes = [p for p in db if _parse_data(p.get("data", "")).startswith(prefixo_mes)]
    total_mes = len(db_mes)

    # Métricas gerais (todas as propostas, não filtrado por mês)
    total = len(db)
    receita_total = sum(i["valor"] for i in todos_itens_aprov)
    fechou_count = sum(1 for p in db if p.get("status") in ("Fechou", "Fechou Parcial"))
    taxa_geral = (fechou_count / total * 100) if total > 0 else 0

    # Métricas do MÊS selecionado
    itens_aprov_mes = len(itens_mes)
    fechou_mes = sum(1 for p in db_mes if p.get("status") in ("Fechou", "Fechou Parcial"))
    taxa_mes = (fechou_mes / total_mes * 100) if total_mes > 0 else 0
    ticket_medio_mes = (receita_mes / itens_aprov_mes) if itens_aprov_mes > 0 else 0

    # Mês anterior para comparação (delta)
    if mes_sel == 1:
        mes_ant, ano_ant = 12, ano_sel - 1
    else:
        mes_ant, ano_ant = mes_sel - 1, ano_sel
    prefixo_ant = f"{ano_ant}-{mes_ant:02d}"
    itens_ant = [i for i in todos_itens_aprov if i["data_aprovacao"].startswith(prefixo_ant)]
    receita_ant = sum(i["valor"] for i in itens_ant)
    db_ant = [p for p in db if _parse_data(p.get("data", "")).startswith(prefixo_ant)]
    fechou_ant = sum(1 for p in db_ant if p.get("status") in ("Fechou", "Fechou Parcial"))
    taxa_ant = (fechou_ant / len(db_ant) * 100) if len(db_ant) > 0 else 0
    ticket_ant = (receita_ant / len(itens_ant)) if len(itens_ant) > 0 else 0

    def _delta_html(atual, anterior, fmt="valor"):
        """Gera HTML de delta comparativo com mês anterior"""
        if anterior == 0:
            return '<span class="delta-neutral">—</span>'
        diff = atual - anterior
        pct = ((atual - anterior) / anterior * 100) if anterior != 0 else 0
        if fmt == "pct":
            txt = f"{abs(diff):.1f}pp"
        elif fmt == "valor":
            txt = fc(abs(diff))
        else:
            txt = f"{abs(diff):.0f}"
        if diff > 0:
            return f'<span class="delta-up">▲ +{txt}</span>'
        elif diff < 0:
            return f'<span class="delta-down">▼ -{txt}</span>'
        return '<span class="delta-neutral">● igual</span>'

    # ===== MELHOR CLIENTE DO PERÍODO =====
    clientes_mes = {}
    for item in itens_mes:
        c = item.get("cliente", "-")
        if c not in clientes_mes:
            clientes_mes[c] = {"receita": 0, "itens": 0}
        clientes_mes[c]["receita"] += item["valor"]
        clientes_mes[c]["itens"] += 1
    melhor_cliente = max(clientes_mes.items(), key=lambda x: x[1]["receita"]) if clientes_mes else ("-", {"receita": 0, "itens": 0})

    # ===== SERVIÇO MAIS VENDIDO DO PERÍODO =====
    servicos_mes = {}
    for item in itens_mes:
        desc = item.get("descricao", "-").strip()
        if desc and desc != "-":
            # Normalizar nome do serviço (primeira palavra significativa)
            desc_norm = desc[:50]
            if desc_norm not in servicos_mes:
                servicos_mes[desc_norm] = {"count": 0, "receita": 0}
            servicos_mes[desc_norm]["count"] += 1
            servicos_mes[desc_norm]["receita"] += item["valor"]
    melhor_servico = max(servicos_mes.items(), key=lambda x: x[1]["receita"]) if servicos_mes else ("-", {"count": 0, "receita": 0})

    # ===== MÉTRICAS GERAIS (4 cards com delta) =====
    st.markdown(f'<div class="section-title">📊 Visão Geral — {meses_nomes[mes_sel-1]} {ano_sel}</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)

    with c1:
        delta_prop = _delta_html(total_mes, len(db_ant), "int")
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">📋</div>
            <div class="metric-label">Propostas no Mês</div>
            <div class="metric-value">{total_mes}</div>
            <div style="margin-top:4px;">{delta_prop} <span style="font-size:11px;color:#9ca3af;">vs mês anterior</span></div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        delta_fecha = _delta_html(fechou_mes, fechou_ant, "int")
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">✅</div>
            <div class="metric-label">Fechamentos</div>
            <div class="metric-value">{fechou_mes}</div>
            <div style="margin-top:4px;">{delta_fecha} <span style="font-size:11px;color:#9ca3af;">vs mês anterior</span></div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        delta_taxa = _delta_html(taxa_mes, taxa_ant, "pct")
        cor_taxa = "#10b981" if taxa_mes >= 50 else "#f59e0b" if taxa_mes >= 30 else "#ef4444"
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">📈</div>
            <div class="metric-label">Taxa de Conversão</div>
            <div class="metric-value" style="color:{cor_taxa};">{taxa_mes:.1f}%</div>
            <div style="margin-top:4px;">{delta_taxa} <span style="font-size:11px;color:#9ca3af;">vs mês anterior</span></div>
        </div>
        """, unsafe_allow_html=True)
    with c4:
        delta_rec = _delta_html(receita_mes, receita_ant)
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">💰</div>
            <div class="metric-label">Receita Aprovada</div>
            <div class="metric-value">{fc(receita_mes)}</div>
            <div style="margin-top:4px;">{delta_rec} <span style="font-size:11px;color:#9ca3af;">vs mês anterior</span></div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("")

    # ===== META + DESTAQUES DO PERÍODO =====
    col_meta, col_destaques = st.columns([3, 2])

    with col_meta:
        st.markdown(f'<div class="section-title">🎯 Meta de Vendas — {meses_nomes[mes_sel-1]} {ano_sel}</div>', unsafe_allow_html=True)
        progresso = (receita_mes / meta_mensal * 100) if meta_mensal > 0 else 0
        progresso_cap = min(progresso, 100)
        falta = max(meta_mensal - receita_mes, 0)

        cache_key = f"meta_celebrada_{prefixo_mes}"
        if progresso < 100 and cache_key in st.session_state:
            del st.session_state[cache_key]

        if progresso >= 100:
            bar_class = "progress-green"
            emoji_status = "🏆"
            txt_status = "META BATIDA!"
        elif progresso >= 60:
            bar_class = "progress-gold"
            emoji_status = "🔥"
            txt_status = "Bom ritmo!"
        else:
            bar_class = "progress-red"
            emoji_status = "⚡"
            txt_status = "Vamos acelerar!"

        st.markdown(f"""
        <div class="meta-card">
            <div style="display:flex; justify-content:space-between; align-items:flex-start;">
                <div>
                    <div class="meta-title">Progresso da Meta</div>
                    <div class="meta-valor">{fc(receita_mes)}</div>
                    <div class="meta-sub">de {fc(meta_mensal)}</div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:40px;">{emoji_status}</div>
                    <div style="font-weight:700; color:#1a2744; font-size:22px;">{progresso:.0f}%</div>
                    <div style="font-size:12px; color:#6b7280;">{txt_status}</div>
                </div>
            </div>
            <div class="progress-container">
                <div class="progress-bar {bar_class}" style="width:{progresso_cap}%;"></div>
            </div>
            <div style="display:flex; justify-content:space-between; font-size:12px; color:#6b7280;">
                <span>{total_mes} propostas criadas | {itens_aprov_mes} itens aprovados</span>
                <span>Falta: <strong style="color:#1a2744;">{fc(falta)}</strong></span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if progresso >= 100 and cache_key not in st.session_state:
            st.session_state[cache_key] = True
            st.balloons()

    with col_destaques:
        st.markdown(f'<div class="section-title">⭐ Destaques do Período</div>', unsafe_allow_html=True)
        # Melhor Cliente
        st.markdown(f"""
        <div class="highlight-card hl-gold" style="margin-bottom:12px;">
            <div class="hl-icon hl-icon-gold">🏆</div>
            <div class="hl-label">Melhor Cliente</div>
            <div class="hl-value">{melhor_cliente[0]}</div>
            <div class="hl-sub"><strong>{fc(melhor_cliente[1]['receita'])}</strong> · {melhor_cliente[1]['itens']} item(ns)</div>
        </div>
        """, unsafe_allow_html=True)
        # Serviço Mais Vendido
        st.markdown(f"""
        <div class="highlight-card hl-green" style="margin-bottom:12px;">
            <div class="hl-icon hl-icon-green">⭐</div>
            <div class="hl-label">Serviço Mais Vendido</div>
            <div class="hl-value" title="{melhor_servico[0]}">{melhor_servico[0][:35]}{'...' if len(melhor_servico[0]) > 35 else ''}</div>
            <div class="hl-sub"><strong>{fc(melhor_servico[1]['receita'])}</strong> · {melhor_servico[1]['count']}x vendido</div>
        </div>
        """, unsafe_allow_html=True)
        # Ticket Médio
        delta_ticket = _delta_html(ticket_medio_mes, ticket_ant)
        st.markdown(f"""
        <div class="highlight-card hl-purple">
            <div class="hl-icon hl-icon-purple">💎</div>
            <div class="hl-label">Ticket Médio</div>
            <div class="hl-value">{fc(ticket_medio_mes)}</div>
            <div class="hl-sub">{delta_ticket} <span style="font-size:11px;color:#9ca3af;">vs mês anterior</span></div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("")

    # ===== RANKING DE VENDEDORES (baseado na data_aprovacao do mês) =====
    col_rank, col_status = st.columns([3, 2])

    with col_rank:
        st.markdown(f'<div class="section-title">🏅 Ranking de Vendedores — {meses_nomes[mes_sel-1]} {ano_sel}</div>', unsafe_allow_html=True)
        fotos_vendedores = config.get("vendedores_fotos", {})
        vendedores_stats = {}
        for item in itens_mes:
            v = item["vendedor"]
            if v not in vendedores_stats:
                vendedores_stats[v] = {"receita": 0, "itens": 0}
            vendedores_stats[v]["receita"] += item["valor"]
            vendedores_stats[v]["itens"] += 1

        # Contagem de propostas por vendedor no mês
        propostas_por_vendedor = {}
        for p in db_mes:
            v = p.get("vendedor", "Sem vendedor") or "Sem vendedor"
            propostas_por_vendedor[v] = propostas_por_vendedor.get(v, 0) + 1

        ranking = sorted(vendedores_stats.items(), key=lambda x: x[1]["receita"], reverse=True)

        if ranking:
            for pos, (nome, stats) in enumerate(ranking[:5], 1):
                pos_class = f"ranking-{pos}" if pos <= 3 else "ranking-other"
                medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(pos, f"{pos}º")
                n_propostas = propostas_por_vendedor.get(nome, 0)
                avatar = get_avatar_html(nome, fotos_vendedores)
                st.markdown(f"""
                <div class="ranking-row">
                    <div class="ranking-pos {pos_class}">{medal if pos <= 3 else pos}</div>
                    {avatar}
                    <div class="ranking-name">{nome}</div>
                    <div class="ranking-stats">
                        {stats['itens']} item(ns) aprovado(s) · {n_propostas} proposta(s)<br>
                        <span class="ranking-valor">{fc(stats['receita'])}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("Nenhum item aprovado neste mês.")

    with col_status:
        st.markdown(f'<div class="section-title">🔄 Funil de Conversão — {meses_nomes[mes_sel-1]}</div>', unsafe_allow_html=True)
        # Funil baseado no mês selecionado
        counts_mes = {"Enviada": 0, "Pendente": 0, "Fechou": 0, "Fechou Parcial": 0, "Não Fechou": 0}
        for p in db_mes:
            s = p.get("status", "Enviada")
            if s in counts_mes:
                counts_mes[s] += 1

        # Funnel steps ordered by pipeline stage
        funnel_data = [
            ("Propostas", total_mes, "#6b7280"),
            ("Enviadas", counts_mes["Enviada"], "#3b82f6"),
            ("Pendentes", counts_mes["Pendente"], "#f59e0b"),
            ("Fechou", counts_mes["Fechou"] + counts_mes["Fechou Parcial"], "#10b981"),
            ("Não Fechou", counts_mes["Não Fechou"], "#ef4444"),
        ]
        max_funnel = max(total_mes, 1)
        for label, count, color in funnel_data:
            pct = (count / max_funnel * 100) if max_funnel > 0 else 0
            pct_display = max(pct, 8) if count > 0 else 0
            st.markdown(f"""
            <div class="funnel-step">
                <div class="funnel-label">{label}</div>
                <div class="funnel-bar-bg">
                    <div class="funnel-bar" style="width:{pct_display}%; background:{color};">{count}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        # Insight automático
        if total_mes > 0:
            if taxa_mes >= 60:
                insight_txt = f"Excelente! Taxa de <strong>{taxa_mes:.0f}%</strong> de conversão. O time está performando muito bem em {meses_nomes[mes_sel-1]}."
                insight_emoji = "🚀"
            elif taxa_mes >= 35:
                insight_txt = f"Bom desempenho com <strong>{taxa_mes:.0f}%</strong> de conversão. Ainda há espaço para melhorar o follow-up."
                insight_emoji = "💡"
            elif taxa_mes > 0:
                nao_fechou = counts_mes["Não Fechou"]
                if nao_fechou > fechou_mes:
                    insight_txt = f"Atenção: <strong>{nao_fechou}</strong> propostas não fecharam vs <strong>{fechou_mes}</strong> fechadas. Revise a abordagem comercial."
                    insight_emoji = "⚠️"
                else:
                    insight_txt = f"Taxa de <strong>{taxa_mes:.0f}%</strong>. Foque no follow-up das <strong>{counts_mes['Enviada'] + counts_mes['Pendente']}</strong> propostas ainda abertas."
                    insight_emoji = "📌"
            else:
                insight_txt = f"Nenhum fechamento ainda em {meses_nomes[mes_sel-1]}. Hora de acelerar o follow-up das <strong>{total_mes}</strong> propostas enviadas!"
                insight_emoji = "⚡"
            st.markdown(f"""
            <div class="insight-box">
                <p>{insight_emoji} {insight_txt}</p>
            </div>
            """, unsafe_allow_html=True)

    # ===== TOP CLIENTES DO MÊS (tabela) =====
    if clientes_mes:
        st.markdown(f'<div class="section-title">🏅 Top Clientes — {meses_nomes[mes_sel-1]} {ano_sel}</div>', unsafe_allow_html=True)
        top_clientes = sorted(clientes_mes.items(), key=lambda x: x[1]["receita"], reverse=True)[:5]
        for pos, (nome_cli, stats_cli) in enumerate(top_clientes, 1):
            medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(pos, f"  {pos}º")
            pct_meta = (stats_cli['receita'] / meta_mensal * 100) if meta_mensal > 0 else 0
            st.markdown(f"""
            <div style="display:flex; align-items:center; padding:10px 16px; background:white; border-radius:10px; margin-bottom:6px; border:1px solid #f0f0f0;">
                <div style="width:32px; text-align:center; font-size:16px;">{medal}</div>
                <div style="flex:1; margin-left:12px;">
                    <strong style="color:#1a2744; font-size:14px;">{nome_cli}</strong>
                    <span style="color:#9ca3af; font-size:11px; margin-left:8px;">{stats_cli['itens']} serviço(s)</span>
                </div>
                <div style="text-align:right;">
                    <div style="font-weight:700; color:#1a2744; font-size:14px;">{fc(stats_cli['receita'])}</div>
                    <div style="font-size:11px; color:#9ca3af;">{pct_meta:.1f}% da meta</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("")

    # ===== ÚLTIMAS PROPOSTAS =====
    if db:
        st.markdown('<div class="section-title">🕐 Últimas Propostas</div>', unsafe_allow_html=True)
        for p in db[:5]:
            status = p.get("status", "Enviada")
            badge_class = {"Enviada": "status-enviada", "Fechou": "status-fechou", "Fechou Parcial": "status-fechou-parcial", "Não Fechou": "status-nao-fechou", "Pendente": "status-pendente"}.get(status, "status-enviada")
            vendedor_p = p.get("vendedor", "-")
            st.markdown(f"""
            <div style="display:flex; align-items:center; padding:12px 16px; background:white; border-radius:10px; margin-bottom:8px; border:1px solid #f0f0f0;">
                <div style="flex:1;">
                    <strong style="color:#1a2744;">{p.get('cliente', '-')}</strong>
                    <span style="color:#9ca3af; font-size:12px; margin-left:8px;">{p.get('data', '')}</span>
                    <span style="color:#6b7280; font-size:11px; margin-left:8px;">· {vendedor_p}</span>
                </div>
                <div style="margin-right:16px; font-weight:700; color:#1a2744;">{fc(p.get('valor', 0))}</div>
                <span class="status-badge {badge_class}">{status}</span>
            </div>
            """, unsafe_allow_html=True)


# ==========================================
# TAB: NOVA PROPOSTA
# ==========================================
with tab_nova:
    config = load_config()
    vendedores_lista = config.get("vendedores", ["Allan"])

    with st.form("proposta_form"):
        st.markdown("#### 👤 Dados do Cliente")
        col1, col2 = st.columns(2)
        with col1:
            tratamento = st.selectbox("Tratamento", ["Ao Sr.", "À Sra.", "A", "Ao", "Prezado(a)"])
            telefone = st.text_input("Telefone", placeholder="(84) 99999-0000")
        with col2:
            nome_cliente = st.text_input("Nome do Cliente *", placeholder="Ex: João da Silva")
            email_cliente = st.text_input("E-mail", placeholder="cliente@email.com")
        nome_empresa = st.text_input("🏢 Nome da Empresa (opcional)", placeholder="Ex: Empresa XYZ LTDA — deixe em branco se ainda vai abrir")

        st.markdown("#### 👨‍💼 Vendedor")
        vendedor = st.selectbox("Vendedor responsável", vendedores_lista)

        st.markdown("#### 📝 Descrição da Proposta")
        introducao = st.text_area(
            "Introdução (referente a...)",
            placeholder="Ex: gestão contábil e fiscal da empresa NOME LTDA",
            height=68
        )

        st.markdown("#### 💰 Serviços e Honorários")
        col_desc, col_per = st.columns(2)
        with col_desc:
            desconto_pct = st.selectbox("Desconto fictício", [
                ("10%", 0.10), ("15%", 0.15), ("20%", 0.20), ("Sem desconto", 0.0)
            ], format_func=lambda x: x[0])
        with col_per:
            periodicidade_padrao = st.selectbox("Periodicidade padrão", ["Mensal", "Única Vez", "Trimestral", "Anual"])

        st.markdown("**Adicione os serviços:**")
        num_svcs = st.number_input("Quantidade de serviços", min_value=1, max_value=10, value=1)

        servicos = []
        for i in range(int(num_svcs)):
            st.markdown(f"**Serviço {i+1}**")
            cs1, cs2, cs3 = st.columns([3, 1, 1])
            with cs1:
                desc_svc = st.text_input(f"Descrição", key=f"svc_desc_{i}", placeholder="Ex: Gestão contábil mensal")
            with cs2:
                valor_svc = st.text_input(f"Valor Real (R$)", key=f"svc_val_{i}", placeholder="500,00")
            with cs3:
                per_svc = st.selectbox(f"Periodicidade", ["Mensal", "Única Vez", "Trimestral", "Anual"], key=f"svc_per_{i}",
                                       index=["Mensal", "Única Vez", "Trimestral", "Anual"].index(periodicidade_padrao))
            servicos.append({"descricao": desc_svc, "valor": valor_svc, "periodicidade": per_svc})

        st.markdown("#### 💳 Pagamento e Extras")
        pix_opcao = st.selectbox("Chave PIX para pagamento", [
            ("PF — 33.540.066/0001-23 (Allan Sayure)", "33.540.066/0001-23", "ALLAN SAYURE DE AZEVEDO BARBOSA"),
            ("PJ — 35.304.872/0001-28 (Azevedo Contabilidade)", "35.304.872/0001-28", "AZEVEDO CONTABILIDADE LTDA")
        ], format_func=lambda x: x[0])

        observacao = st.text_area("Observação (opcional)", placeholder="Ex: O valor refere-se exclusivamente a...", height=68)

        incluir_doc = st.checkbox("Incluir seção de Documentação Necessária")
        texto_doc = ""
        if incluir_doc:
            texto_doc = st.text_area(
                "Texto da documentação",
                value="Cópia da identidade do responsável de cada empresa para fazermos a procuração necessária e assim enviamos para assinatura pelo Gov.br ou certificado digital.",
                height=68
            )

        obs_internas = st.text_area("📌 Observações internas (não aparece na proposta)", placeholder="Anotações internas...", height=68)

        submitted = st.form_submit_button("📄 Gerar Proposta DOCX", use_container_width=True, type="primary")

    if submitted:
        if not nome_cliente.strip():
            st.error("Preencha o nome do cliente!")
        elif not any(s["descricao"].strip() for s in servicos):
            st.error("Adicione pelo menos um serviço com descrição!")
        else:
            def parse_valor(v):
                v = v.replace(".", "").replace(",", ".").strip()
                try:
                    return float(v)
                except:
                    return 0.0

            svcs_parsed = []
            for s in servicos:
                if s["descricao"].strip():
                    svcs_parsed.append({
                        "descricao": s["descricao"],
                        "valor": parse_valor(s["valor"]),
                        "periodicidade": s["periodicidade"]
                    })

            dados = {
                "tratamento": tratamento,
                "nome": nome_cliente,
                "telefone": telefone,
                "email": email_cliente,
                "empresa": nome_empresa,
                "vendedor": vendedor,
                "introducao": introducao or "prestação de serviços contábeis",
                "servicos": svcs_parsed,
                "desconto_pct": desconto_pct[1],
                "pix_cnpj": pix_opcao[1],
                "pix_titular": pix_opcao[2],
                "observacao": observacao,
                "incluir_doc": incluir_doc,
                "texto_doc": texto_doc,
                "logo_path": LOGO_PATH
            }

            with st.spinner("Gerando proposta..."):
                docx_bytes = gerar_docx(dados)

            filename = f"Proposta - {nome_cliente}.docx"
            st.success(f"✅ Proposta gerada com sucesso!")

            # Salvar bytes no session_state para enviar ao Autentique
            st.session_state["_last_docx_bytes"] = docx_bytes
            st.session_state["_last_docx_nome"] = nome_cliente

            col_download, col_autentique = st.columns(2)
            with col_download:
                st.download_button(
                    label="⬇️ Baixar Proposta DOCX",
                    data=docx_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    use_container_width=True
                )
            with col_autentique:
                if st.button("✍️ Enviar para Assinatura (Autentique)", use_container_width=True, type="primary"):
                    with st.spinner("Enviando para Autentique..."):
                        ok, resultado = enviar_para_autentique(
                            docx_bytes,
                            f"Proposta - {nome_cliente}",
                            nome_cliente
                        )
                    if ok:
                        st.session_state["_autentique_resultado"] = resultado
                        st.success("✅ Proposta enviada para assinatura!")
                        link = resultado.get("autentique_link", "")
                        if link:
                            st.info(f"🔗 **Link para o cliente aprovar:**")
                            st.code(link, language=None)
                            st.caption("Copie o link acima e envie ao cliente para aprovação.")
                    else:
                        st.error(f"❌ Erro ao enviar: {resultado}")

            total_valor = sum(s["valor"] for s in svcs_parsed)

            # Serviços detalhados com status individual (Pendente por padrão)
            svcs_detalhados = []
            for s in svcs_parsed:
                svcs_detalhados.append({
                    "descricao": s["descricao"],
                    "valor": s["valor"],
                    "periodicidade": s["periodicidade"],
                    "status": "Pendente"
                })

            nova_proposta = {
                "id": int(datetime.now().timestamp() * 1000),
                "data": datetime.now().strftime("%Y-%m-%d"),
                "cliente": nome_cliente,
                "empresa": nome_empresa,
                "tratamento": tratamento,
                "telefone": telefone,
                "email": email_cliente,
                "vendedor": vendedor,
                "servicos": "; ".join(s["descricao"] for s in svcs_parsed),
                "valor": total_valor,
                "status": "Enviada",
                "obs": obs_internas,
                "servicos_detalhados": json.dumps(svcs_detalhados, ensure_ascii=False),
                "autentique_id": "",
                "autentique_link": ""
            }

            if USING_SHEETS:
                save_proposta(nova_proposta)
            else:
                db = load_db()
                db.insert(0, nova_proposta)
                save_db(db)


# ==========================================
# TAB: HISTÓRICO
# ==========================================
with tab_hist:
    db = load_db()

    # Executar expiração automática de itens pendentes > 30 dias
    if not st.session_state.get("_expirado_check_done"):
        if USING_SHEETS:
            expirar_itens_pendentes(30)
        st.session_state["_expirado_check_done"] = True

    col_search, col_filter, col_export = st.columns([3, 2, 1])
    with col_search:
        busca = st.text_input("🔍 Buscar cliente ou serviço", placeholder="Digite para filtrar...")
    with col_filter:
        filtro_status = st.selectbox("Filtrar por status", ["Todos", "Enviada", "Fechou", "Fechou Parcial", "Não Fechou", "Pendente"])
    with col_export:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📥 Exportar CSV", use_container_width=True):
            if db:
                import csv
                import io
                output = io.StringIO()
                writer = csv.writer(output, delimiter=";")
                writer.writerow(["Data", "Cliente", "Serviços", "Valor Aprovado", "Valor Total", "Vendedor", "Status", "Observações"])
                for p in db:
                    # Calcular valor aprovado dos itens
                    val_aprovado = 0
                    val_total = p.get("valor", 0)
                    sd = p.get("servicos_detalhados", "")
                    if sd and sd.strip().startswith("["):
                        try:
                            svcs = json.loads(sd)
                            val_aprovado = sum(s.get("valor", 0) for s in svcs if s.get("status") == "Aprovado")
                            val_total = sum(s.get("valor", 0) for s in svcs)
                        except Exception:
                            val_aprovado = val_total if p.get("status") in ("Fechou", "Fechou Parcial") else 0
                    else:
                        # Proposta antiga — valor integral se fechou
                        if p.get("status") in ("Fechou", "Fechou Parcial"):
                            val_aprovado = val_total
                    writer.writerow([
                        p.get("data", ""), p.get("cliente", ""), p.get("servicos", ""),
                        str(val_aprovado).replace(".", ","),
                        str(val_total).replace(".", ","),
                        p.get("vendedor", ""), p.get("status", ""), p.get("obs", "")
                    ])
                csv_data = "\ufeff" + output.getvalue()
                st.download_button("⬇️ Baixar CSV", csv_data.encode("utf-8"), "propostas.csv", "text/csv")

    filtered = db
    if busca:
        busca_lower = busca.lower()
        filtered = [p for p in filtered if busca_lower in p.get("cliente", "").lower() or busca_lower in p.get("servicos", "").lower()]
    if filtro_status != "Todos":
        filtered = [p for p in filtered if p.get("status") == filtro_status]

    if not filtered:
        st.info("Nenhuma proposta encontrada.")
    else:
        for i, p in enumerate(filtered):
            status = p.get("status", "Enviada")
            status_color = {"Enviada": "🔵", "Fechou": "🟢", "Fechou Parcial": "🔵🟢", "Não Fechou": "🔴", "Pendente": "🟡"}.get(status, "⚪")

            # Calcular valor aprovado para exibição
            sd_str = p.get("servicos_detalhados", "")
            valor_aprovado = 0
            valor_total = p.get("valor", 0)
            if sd_str and sd_str.strip().startswith("["):
                try:
                    svcs_det = json.loads(sd_str)
                    valor_aprovado = sum(s.get("valor", 0) for s in svcs_det if s.get("status") == "Aprovado")
                    valor_total_calc = sum(s.get("valor", 0) for s in svcs_det)
                    if valor_total_calc > 0:
                        valor_total = valor_total_calc
                except Exception:
                    valor_aprovado = valor_total if status in ("Fechou", "Fechou Parcial") else 0
            else:
                # Proposta antiga — valor integral se fechou
                if status in ("Fechou", "Fechou Parcial"):
                    valor_aprovado = valor_total

            valor_display = fc(valor_aprovado) if valor_aprovado > 0 else fc(valor_total)
            label_extra = f" (aprovado: {fc(valor_aprovado)})" if valor_aprovado > 0 and valor_aprovado != valor_total else ""

            with st.expander(f"{status_color} **{p['cliente']}** — {p.get('data', '')} — {fc(valor_total)}{label_extra} — {status}"):
                c1, c2 = st.columns(2)
                with c1:
                    st.write(f"**Vendedor:** {p.get('vendedor', '-')}")
                    st.write(f"**Telefone:** {p.get('telefone', '-')}")
                    empresa_hist = p.get('empresa', '')
                    if empresa_hist and str(empresa_hist).strip() and str(empresa_hist) not in ('nan', 'None', ''):
                        st.write(f"**Empresa:** {empresa_hist}")
                with c2:
                    st.write(f"**E-mail:** {p.get('email', '-')}")
                    st.write(f"**Data:** {p.get('data', '-')}")

                # ===== SERVIÇOS COM STATUS INDIVIDUAL =====
                st.markdown("---")
                st.markdown("#### 📦 Serviços da Proposta")

                if sd_str:
                    try:
                        svcs_det = json.loads(sd_str)
                    except Exception:
                        svcs_det = []
                else:
                    # Proposta antiga sem servicos_detalhados — criar a partir do campo servicos
                    svcs_det = []
                    servicos_texto = p.get("servicos", "")
                    if servicos_texto:
                        for desc in servicos_texto.split(";"):
                            desc = desc.strip()
                            if desc:
                                svcs_det.append({
                                    "descricao": desc,
                                    "valor": 0,
                                    "periodicidade": "-",
                                    "status": "Pendente" if status == "Enviada" else ("Aprovado" if status == "Fechou" else "Recusado")
                                })

                if svcs_det:
                    alterou_servicos = False
                    for si, svc in enumerate(svcs_det):
                        svc_status = svc.get("status", "Pendente")

                        col_desc, col_val, col_per, col_st, col_dt = st.columns([2.5, 1.2, 0.8, 1.2, 1.3])
                        with col_desc:
                            st.markdown(f"**{svc.get('descricao', '-')}**")
                        with col_val:
                            st.markdown(f"**{fc(svc.get('valor', 0))}**")
                        with col_per:
                            st.markdown(f"*{svc.get('periodicidade', '-')}*")
                        with col_st:
                            opcoes_status = ["Pendente", "Aprovado", "Recusado"]
                            idx_atual = opcoes_status.index(svc_status) if svc_status in opcoes_status else 0
                            novo_svc_status = st.selectbox(
                                f"Status",
                                opcoes_status,
                                index=idx_atual,
                                key=f"svc_st_{p['id']}_{si}",
                                label_visibility="collapsed"
                            )
                            if novo_svc_status != svc_status:
                                svcs_det[si]["status"] = novo_svc_status
                                alterou_servicos = True
                        with col_dt:
                            if novo_svc_status == "Aprovado" or svc_status == "Aprovado":
                                # Data de aprovação: pré-preenchida com hoje, editável, só passado
                                data_aprov_atual = svc.get("data_aprovacao", "")
                                try:
                                    default_date = datetime.strptime(data_aprov_atual, "%Y-%m-%d").date() if data_aprov_atual else date.today()
                                except Exception:
                                    default_date = date.today()
                                nova_data = st.date_input(
                                    "Aprovado em",
                                    value=default_date,
                                    max_value=date.today(),
                                    key=f"svc_dt_{p['id']}_{si}",
                                    label_visibility="collapsed"
                                )
                                nova_data_str = nova_data.strftime("%Y-%m-%d")
                                if nova_data_str != data_aprov_atual:
                                    svcs_det[si]["data_aprovacao"] = nova_data_str
                                    alterou_servicos = True
                            else:
                                st.markdown("<span style='color:#9ca3af; font-size:11px;'>—</span>", unsafe_allow_html=True)

                    if alterou_servicos:
                        # Registrar data de aprovação nos itens aprovados
                        hoje = datetime.now().strftime("%Y-%m-%d")
                        for s in svcs_det:
                            if s.get("status") == "Aprovado" and not s.get("data_aprovacao"):
                                s["data_aprovacao"] = hoje
                            elif s.get("status") != "Aprovado":
                                s.pop("data_aprovacao", None)

                        novo_json = json.dumps(svcs_det, ensure_ascii=False)
                        if USING_SHEETS:
                            ok, msg = update_servicos_detalhados(p["id"], novo_json)
                            if not ok:
                                st.error(f"Erro ao salvar status do serviço: {msg}")
                        else:
                            for item in db:
                                if item["id"] == p["id"]:
                                    item["servicos_detalhados"] = novo_json
                                    # Recalcular status geral
                                    statuses = [s.get("status", "Pendente") for s in svcs_det]
                                    if all(s == "Aprovado" for s in statuses):
                                        item["status"] = "Fechou"
                                    elif all(s in ("Recusado", "Expirado") for s in statuses):
                                        item["status"] = "Não Fechou"
                                    elif any(s == "Aprovado" for s in statuses):
                                        item["status"] = "Fechou Parcial"
                                    else:
                                        item["status"] = "Enviada"
                                    # Valor = só aprovados
                                    item["valor"] = sum(s.get("valor", 0) for s in svcs_det if s.get("status") == "Aprovado")
                                    break
                            save_db(db)
                        st.rerun()

                    # Resumo visual
                    qtd_aprov = sum(1 for s in svcs_det if s.get("status") == "Aprovado")
                    qtd_recus = sum(1 for s in svcs_det if s.get("status") == "Recusado")
                    qtd_pend = sum(1 for s in svcs_det if s.get("status") == "Pendente")
                    qtd_exp = sum(1 for s in svcs_det if s.get("status") == "Expirado")
                    st.markdown(f"""
                    <div style="display:flex; gap:16px; margin-top:8px; font-size:12px;">
                        <span class="svc-status-aprovado">✅ {qtd_aprov} aprovado(s)</span>
                        <span class="svc-status-recusado">❌ {qtd_recus} recusado(s)</span>
                        <span class="svc-status-pendente">⏳ {qtd_pend} pendente(s)</span>
                        {"<span class='svc-status-expirado'>⌛ " + str(qtd_exp) + " expirado(s)</span>" if qtd_exp > 0 else ""}
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.write(f"**Serviços:** {p.get('servicos', '-')}")

                st.markdown("---")

                # Exibir observações gerais
                if p.get("obs"):
                    st.write(f"**Observações:** {p['obs']}")

                # Exibir motivo da perda e histórico
                if status in ("Não Fechou",):
                    if p.get("motivo_perda"):
                        st.markdown(f"🔴 **Último motivo:** {p['motivo_perda']}")

                    # Botão para adicionar/atualizar motivo
                    if st.session_state.get(f"add_motivo_{p['id']}", False):
                        motivo_add = st.text_area(
                            "Descreva o motivo da perda",
                            placeholder="Ex: Preço alto, foi para concorrente, não respondeu mais...",
                            key=f"motivo_add_{p['id']}"
                        )
                        col_s, col_c = st.columns(2)
                        with col_s:
                            if st.button("✅ Salvar", key=f"salvar_add_{p['id']}"):
                                if motivo_add.strip():
                                    historico_ant = str(p.get("historico", "") or "")
                                    if USING_SHEETS:
                                        ok, msg = update_proposta_status(p["id"], "Não Fechou", motivo_add.strip(), historico_ant)
                                        if not ok:
                                            st.error(f"Erro ao salvar motivo: {msg}")
                                        else:
                                            st.session_state.pop(f"add_motivo_{p['id']}", None)
                                            st.rerun()
                                    else:
                                        for item in db:
                                            if item["id"] == p["id"]:
                                                item["motivo_perda"] = motivo_add.strip()
                                                data_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
                                                nova_entrada = f"[{data_hora}] {motivo_add.strip()}"
                                                item["historico"] = f"{historico_ant} | {nova_entrada}" if historico_ant else nova_entrada
                                                break
                                        save_db(db)
                                        st.session_state.pop(f"add_motivo_{p['id']}", None)
                                        st.rerun()
                                else:
                                    st.warning("Preencha o motivo.")
                        with col_c:
                            if st.button("❌ Cancelar", key=f"cancelar_add_{p['id']}"):
                                st.session_state.pop(f"add_motivo_{p['id']}", None)
                                st.rerun()
                    else:
                        if st.button("📝 Adicionar/Atualizar motivo", key=f"btn_add_motivo_{p['id']}"):
                            st.session_state[f"add_motivo_{p['id']}"] = True
                            st.rerun()

                if p.get("historico"):
                    with st.expander("📋 Histórico de observações"):
                        entradas = str(p["historico"]).split(" | ")
                        for entrada in entradas:
                            if entrada.strip():
                                st.markdown(f"- {entrada}")

                # ===== AUTENTIQUE: Assinatura digital =====
                autentique_id_atual = str(p.get("autentique_id", "") or "").strip()
                autentique_link_atual = str(p.get("autentique_link", "") or "").strip()
                # Filtrar valores "nan" vindos do Sheets
                if autentique_id_atual.lower() in ("nan", "none"):
                    autentique_id_atual = ""
                if autentique_link_atual.lower() in ("nan", "none"):
                    autentique_link_atual = ""

                col_aut1, col_aut2 = st.columns(2)
                with col_aut1:
                    if not autentique_id_atual:
                        if st.button("✍️ Enviar para Assinatura", key=f"aut_send_{p['id']}", use_container_width=True):
                            try:
                                # Reconstruir dados mínimos para gerar o DOCX
                                svcs_para_docx = []
                                try:
                                    sd = json.loads(p.get("servicos_detalhados", "[]") or "[]")
                                    for s in sd:
                                        svcs_para_docx.append({
                                            "descricao": s.get("descricao", ""),
                                            "valor": float(s.get("valor", 0) or 0),
                                            "periodicidade": s.get("periodicidade", "Mensal")
                                        })
                                except Exception:
                                    # Fallback: usa campo "servicos" (texto)
                                    for desc in str(p.get("servicos", "")).split(";"):
                                        if desc.strip():
                                            svcs_para_docx.append({
                                                "descricao": desc.strip(),
                                                "valor": 0.0,
                                                "periodicidade": "Mensal"
                                            })

                                dados_hist = {
                                    "tratamento": p.get("tratamento", "Sr."),
                                    "nome": p.get("cliente", ""),
                                    "telefone": p.get("telefone", ""),
                                    "email": p.get("email", ""),
                                    "empresa": p.get("empresa", ""),
                                    "vendedor": p.get("vendedor", ""),
                                    "introducao": "prestação de serviços contábeis",
                                    "servicos": svcs_para_docx,
                                    "desconto_pct": 0,
                                    "pix_cnpj": "35.304.872/0001-28",
                                    "pix_titular": "AZEVEDO CONTABILIDADE LTDA",
                                    "observacao": "",
                                    "incluir_doc": False,
                                    "texto_doc": "",
                                    "logo_path": LOGO_PATH
                                }

                                with st.spinner("Gerando proposta e enviando ao Autentique..."):
                                    docx_bytes_hist = gerar_docx(dados_hist)
                                    ok_aut, resultado_aut = enviar_para_autentique(
                                        docx_bytes_hist,
                                        f"Proposta - {p.get('cliente', '')}",
                                        p.get("cliente", "Cliente")
                                    )

                                if ok_aut:
                                    aut_id = resultado_aut.get("autentique_id", "")
                                    aut_link = resultado_aut.get("autentique_link", "")
                                    if USING_SHEETS:
                                        ok_save, msg_save = update_proposta_autentique(p["id"], aut_id, aut_link)
                                        if not ok_save:
                                            st.warning(f"Enviado, mas erro ao salvar: {msg_save}")
                                    else:
                                        for item in db:
                                            if item["id"] == p["id"]:
                                                item["autentique_id"] = aut_id
                                                item["autentique_link"] = aut_link
                                                break
                                        save_db(db)
                                    st.success("✅ Enviado para assinatura!")
                                    st.rerun()
                                else:
                                    st.error(f"❌ Erro: {resultado_aut}")
                            except Exception as e:
                                st.error(f"❌ Erro inesperado: {str(e)}")
                    else:
                        st.success("✅ Enviado ao Autentique")

                with col_aut2:
                    if autentique_link_atual:
                        st.markdown("**🔗 Link do cliente:**")
                        st.code(autentique_link_atual, language=None)
                    elif autentique_id_atual:
                        # Documento existe no Autentique mas link ainda não foi salvo
                        if st.button("🔄 Buscar link", key=f"aut_fetch_{p['id']}", use_container_width=True):
                            with st.spinner("Buscando link no Autentique..."):
                                ok_q, doc = consultar_autentique(autentique_id_atual)
                            if ok_q and doc:
                                novo_link = ""
                                for sig in doc.get("signatures", []):
                                    action_name = (sig.get("action") or {}).get("name", "")
                                    if action_name == "APPROVE" or (not sig.get("email")):
                                        link_info = sig.get("link") or {}
                                        sl = link_info.get("short_link", "")
                                        if sl:
                                            novo_link = sl
                                            break
                                if novo_link:
                                    if USING_SHEETS:
                                        update_proposta_autentique(p["id"], autentique_id_atual, novo_link)
                                    else:
                                        for item in db:
                                            if item["id"] == p["id"]:
                                                item["autentique_link"] = novo_link
                                                break
                                        save_db(db)
                                    st.success(f"✅ Link encontrado: {novo_link}")
                                    st.rerun()
                                else:
                                    st.warning("Link ainda não disponível. Tente novamente em alguns segundos.")
                            else:
                                st.error(f"Erro ao consultar: {doc}")

                if st.button(f"🗑️ Excluir", key=f"del_{p['id']}"):
                    if USING_SHEETS:
                        delete_proposta(p["id"])
                    else:
                        db = [item for item in db if item["id"] != p["id"]]
                        save_db(db)
                    st.rerun()


# ==========================================
# TAB: COMISSÕES
# ==========================================
with tab_comissao:
    db_com = load_db()
    config_com = load_config()

    st.markdown("#### 💰 Relatório de Comissões")
    st.caption("Receita item a item — Mensal (Recorrente) vs Avulso — para conferência de comissões")

    # ===== HELPER local: extrair itens com periodicidade =====
    def _parse_data_com(data_str):
        if not data_str:
            return ""
        data_str = str(data_str).strip()
        if len(data_str) >= 10 and data_str[4] == "-":
            return data_str[:10]
        if "/" in data_str:
            parts = data_str.split("/")
            if len(parts) == 3 and len(parts[2]) == 4:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        return data_str

    def extrair_itens_comissao(propostas):
        """Retorna lista de dicts com todos os campos necessários para comissão"""
        itens = []
        for p in propostas:
            sd = p.get("servicos_detalhados", "")
            vendedor = p.get("vendedor", "Sem vendedor") or "Sem vendedor"
            data_proposta = _parse_data_com(p.get("data", ""))
            cliente = p.get("cliente", "-")

            if sd and str(sd).strip().startswith("["):
                try:
                    svcs = json.loads(sd)
                    itens_aprovados = [s for s in svcs if s.get("status") == "Aprovado"]
                    soma_itens = sum(s.get("valor", 0) for s in itens_aprovados)
                    valor_proposta = p.get("valor", 0)

                    if itens_aprovados and soma_itens > 0:
                        for s in itens_aprovados:
                            per = s.get("periodicidade", "-")
                            tipo = "Mensal" if per == "Mensal" else "Avulso"
                            itens.append({
                                "cliente": cliente,
                                "descricao": s.get("descricao", "-"),
                                "valor": s.get("valor", 0),
                                "periodicidade": per,
                                "tipo": tipo,
                                "vendedor": vendedor,
                                "data_aprovacao": _parse_data_com(s.get("data_aprovacao", "")) or data_proposta,
                            })
                    elif itens_aprovados and soma_itens == 0 and valor_proposta > 0:
                        itens.append({
                            "cliente": cliente,
                            "descricao": p.get("servicos", "-"),
                            "valor": valor_proposta,
                            "periodicidade": "-",
                            "tipo": "Avulso",
                            "vendedor": vendedor,
                            "data_aprovacao": _parse_data_com(itens_aprovados[0].get("data_aprovacao", "")) or data_proposta,
                        })
                except Exception:
                    if p.get("status") in ("Fechou", "Fechou Parcial") and p.get("valor", 0) > 0:
                        itens.append({
                            "cliente": cliente,
                            "descricao": p.get("servicos", "-"),
                            "valor": p.get("valor", 0),
                            "periodicidade": "-",
                            "tipo": "Avulso",
                            "vendedor": vendedor,
                            "data_aprovacao": data_proposta,
                        })
            else:
                if p.get("status") in ("Fechou", "Fechou Parcial") and p.get("valor", 0) > 0:
                    itens.append({
                        "cliente": cliente,
                        "descricao": p.get("servicos", "-"),
                        "valor": p.get("valor", 0),
                        "periodicidade": "-",
                        "tipo": "Avulso",
                        "vendedor": vendedor,
                        "data_aprovacao": data_proposta,
                    })
        return itens

    todos_itens_com = extrair_itens_comissao(db_com)

    # ===== FILTROS =====
    st.markdown('<div class="section-title">📅 Filtros</div>', unsafe_allow_html=True)
    col_f1, col_f2, col_f3 = st.columns([1, 1, 1])
    meses_nomes_com = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                       "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    with col_f1:
        mes_com = st.selectbox("Mês", range(1, 13), index=date.today().month - 1,
                               format_func=lambda x: meses_nomes_com[x-1], key="com_mes")
    with col_f2:
        ano_com = st.number_input("Ano", min_value=2024, max_value=2030, value=date.today().year, key="com_ano")
    with col_f3:
        vendedores_com = config_com.get("vendedores", ["Allan"])
        vendedor_filtro = st.selectbox("Vendedor", ["Todos"] + vendedores_com, key="com_vendedor")

    prefixo_com = f"{ano_com}-{mes_com:02d}"

    # Filtrar por mês
    itens_filtrados = [i for i in todos_itens_com if i["data_aprovacao"].startswith(prefixo_com)]

    # Filtrar por vendedor
    if vendedor_filtro != "Todos":
        itens_filtrados = [i for i in itens_filtrados if i["vendedor"] == vendedor_filtro]

    # Separar Mensal vs Avulso
    itens_mensal = [i for i in itens_filtrados if i["tipo"] == "Mensal"]
    itens_avulso = [i for i in itens_filtrados if i["tipo"] == "Avulso"]
    total_mensal = sum(i["valor"] for i in itens_mensal)
    total_avulso = sum(i["valor"] for i in itens_avulso)
    total_geral = total_mensal + total_avulso

    # ===== CARDS DE RESUMO =====
    col_r1, col_r2, col_r3 = st.columns(3)
    with col_r1:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">🔄</div>
            <div class="metric-label">Receita Mensal (Recorrente)</div>
            <div class="metric-value" style="color:#10b981;">{fc(total_mensal)}</div>
            <div style="font-size:12px; color:#9ca3af; margin-top:4px;">{len(itens_mensal)} {'item' if len(itens_mensal) == 1 else 'itens'}</div>
        </div>
        """, unsafe_allow_html=True)
    with col_r2:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">📌</div>
            <div class="metric-label">Receita Avulsa (Única/Trim/Anual)</div>
            <div class="metric-value" style="color:#3b82f6;">{fc(total_avulso)}</div>
            <div style="font-size:12px; color:#9ca3af; margin-top:4px;">{len(itens_avulso)} {'item' if len(itens_avulso) == 1 else 'itens'}</div>
        </div>
        """, unsafe_allow_html=True)
    with col_r3:
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-icon">💰</div>
            <div class="metric-label">Total Fechado no Período</div>
            <div class="metric-value">{fc(total_geral)}</div>
            <div style="font-size:12px; color:#9ca3af; margin-top:4px;">{len(itens_filtrados)} {'item' if len(itens_filtrados) == 1 else 'itens'} aprovados</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ===== TABELA MENSAL =====
    if itens_mensal:
        st.markdown("##### 🔄 Receita Mensal (Recorrente)")
        for idx, item in enumerate(itens_mensal):
            st.markdown(f"""
            <div style="display:flex; align-items:center; padding:10px 16px; background:white; border-radius:10px; margin-bottom:6px; border-left:4px solid #10b981; border:1px solid #f0f0f0;">
                <div style="flex:2;">
                    <strong style="color:#1a2744; font-size:14px;">{item['cliente']}</strong><br>
                    <span style="color:#6b7280; font-size:12px;">{item['descricao']}</span>
                </div>
                <div style="flex:1; text-align:center;">
                    <span style="color:#9ca3af; font-size:11px;">Vendedor</span><br>
                    <span style="font-weight:600; color:#1a2744; font-size:13px;">{item['vendedor']}</span>
                </div>
                <div style="flex:1; text-align:center;">
                    <span style="color:#9ca3af; font-size:11px;">Periodicidade</span><br>
                    <span style="font-weight:500; font-size:13px; color:#10b981;">{item['periodicidade']}</span>
                </div>
                <div style="flex:1; text-align:right;">
                    <span style="font-weight:700; color:#10b981; font-size:15px;">{fc(item['valor'])}</span><br>
                    <span style="color:#9ca3af; font-size:11px;">{item['data_aprovacao']}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown(f"""
        <div style="text-align:right; padding:8px 16px; font-weight:700; color:#10b981; font-size:16px;">
            Subtotal Mensal: {fc(total_mensal)}
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info("Nenhum serviço mensal (recorrente) aprovado neste período.")

    st.markdown("")

    # ===== TABELA AVULSO =====
    if itens_avulso:
        st.markdown("##### 📌 Receita Avulsa")
        for idx, item in enumerate(itens_avulso):
            st.markdown(f"""
            <div style="display:flex; align-items:center; padding:10px 16px; background:white; border-radius:10px; margin-bottom:6px; border-left:4px solid #3b82f6; border:1px solid #f0f0f0;">
                <div style="flex:2;">
                    <strong style="color:#1a2744; font-size:14px;">{item['cliente']}</strong><br>
                    <span style="color:#6b7280; font-size:12px;">{item['descricao']}</span>
                </div>
                <div style="flex:1; text-align:center;">
                    <span style="color:#9ca3af; font-size:11px;">Vendedor</span><br>
                    <span style="font-weight:600; color:#1a2744; font-size:13px;">{item['vendedor']}</span>
                </div>
                <div style="flex:1; text-align:center;">
                    <span style="color:#9ca3af; font-size:11px;">Periodicidade</span><br>
                    <span style="font-weight:500; font-size:13px; color:#3b82f6;">{item['periodicidade']}</span>
                </div>
                <div style="flex:1; text-align:right;">
                    <span style="font-weight:700; color:#3b82f6; font-size:15px;">{fc(item['valor'])}</span><br>
                    <span style="color:#9ca3af; font-size:11px;">{item['data_aprovacao']}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown(f"""
        <div style="text-align:right; padding:8px 16px; font-weight:700; color:#3b82f6; font-size:16px;">
            Subtotal Avulso: {fc(total_avulso)}
        </div>
        """, unsafe_allow_html=True)
    else:
        st.info("Nenhum serviço avulso aprovado neste período.")

    st.markdown("---")

    # ===== RESUMO POR VENDEDOR =====
    if itens_filtrados:
        st.markdown("##### 👥 Resumo por Vendedor")
        vendedores_resumo = {}
        for item in itens_filtrados:
            v = item["vendedor"]
            if v not in vendedores_resumo:
                vendedores_resumo[v] = {"mensal": 0, "avulso": 0, "total": 0, "qtd": 0}
            vendedores_resumo[v]["total"] += item["valor"]
            vendedores_resumo[v]["qtd"] += 1
            if item["tipo"] == "Mensal":
                vendedores_resumo[v]["mensal"] += item["valor"]
            else:
                vendedores_resumo[v]["avulso"] += item["valor"]

        for v, dados in sorted(vendedores_resumo.items(), key=lambda x: x[1]["total"], reverse=True):
            st.markdown(f"""
            <div style="display:flex; align-items:center; padding:14px 20px; background:white; border-radius:12px; margin-bottom:8px; border:1px solid #f0f0f0;">
                <div style="flex:1;">
                    <strong style="color:#1a2744; font-size:15px;">{v}</strong>
                    <span style="color:#9ca3af; font-size:12px; margin-left:8px;">{dados['qtd']} {'item' if dados['qtd'] == 1 else 'itens'}</span>
                </div>
                <div style="text-align:center; margin-right:24px;">
                    <span style="color:#9ca3af; font-size:10px; text-transform:uppercase;">Mensal</span><br>
                    <span style="font-weight:600; color:#10b981;">{fc(dados['mensal'])}</span>
                </div>
                <div style="text-align:center; margin-right:24px;">
                    <span style="color:#9ca3af; font-size:10px; text-transform:uppercase;">Avulso</span><br>
                    <span style="font-weight:600; color:#3b82f6;">{fc(dados['avulso'])}</span>
                </div>
                <div style="text-align:right;">
                    <span style="color:#9ca3af; font-size:10px; text-transform:uppercase;">Total</span><br>
                    <span style="font-weight:700; color:#1a2744; font-size:16px;">{fc(dados['total'])}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("")

    # ===== EXPORTAR EXCEL =====
    if itens_filtrados:
        st.markdown("##### 📥 Exportar para Excel")
        try:
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()

            # --- Aba: Detalhado ---
            ws = wb.active
            ws.title = "Detalhado"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1a2744", end_color="1a2744", fill_type="solid")
            mensal_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
            avulso_fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            money_fmt = '#,##0.00'

            headers = ["Cliente", "Descrição", "Valor (R$)", "Periodicidade", "Tipo", "Vendedor", "Data Aprovação"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row_idx, item in enumerate(itens_filtrados, 2):
                ws.cell(row=row_idx, column=1, value=item["cliente"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=item["descricao"]).border = thin_border
                c_val = ws.cell(row=row_idx, column=3, value=item["valor"])
                c_val.number_format = money_fmt
                c_val.border = thin_border
                ws.cell(row=row_idx, column=4, value=item["periodicidade"]).border = thin_border
                c_tipo = ws.cell(row=row_idx, column=5, value=item["tipo"])
                c_tipo.border = thin_border
                c_tipo.fill = mensal_fill if item["tipo"] == "Mensal" else avulso_fill
                ws.cell(row=row_idx, column=6, value=item["vendedor"]).border = thin_border
                ws.cell(row=row_idx, column=7, value=item["data_aprovacao"]).border = thin_border

            # Linha de total
            last_row = len(itens_filtrados) + 2
            ws.cell(row=last_row, column=2, value="TOTAL").font = Font(bold=True, size=12)
            c_total = ws.cell(row=last_row, column=3, value=total_geral)
            c_total.font = Font(bold=True, size=12)
            c_total.number_format = money_fmt

            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 16

            # --- Aba: Resumo por Vendedor ---
            ws2 = wb.create_sheet("Resumo por Vendedor")
            headers2 = ["Vendedor", "Mensal (R$)", "Avulso (R$)", "Total (R$)", "Qtd Itens"]
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            vendedores_resumo_xl = {}
            for item in itens_filtrados:
                v = item["vendedor"]
                if v not in vendedores_resumo_xl:
                    vendedores_resumo_xl[v] = {"mensal": 0, "avulso": 0, "total": 0, "qtd": 0}
                vendedores_resumo_xl[v]["total"] += item["valor"]
                vendedores_resumo_xl[v]["qtd"] += 1
                if item["tipo"] == "Mensal":
                    vendedores_resumo_xl[v]["mensal"] += item["valor"]
                else:
                    vendedores_resumo_xl[v]["avulso"] += item["valor"]

            for row_idx, (v, d) in enumerate(sorted(vendedores_resumo_xl.items(), key=lambda x: x[1]["total"], reverse=True), 2):
                ws2.cell(row=row_idx, column=1, value=v).border = thin_border
                c = ws2.cell(row=row_idx, column=2, value=d["mensal"])
                c.number_format = money_fmt
                c.border = thin_border
                c = ws2.cell(row=row_idx, column=3, value=d["avulso"])
                c.number_format = money_fmt
                c.border = thin_border
                c = ws2.cell(row=row_idx, column=4, value=d["total"])
                c.number_format = money_fmt
                c.border = thin_border
                c.font = Font(bold=True)
                ws2.cell(row=row_idx, column=5, value=d["qtd"]).border = thin_border

            ws2.column_dimensions['A'].width = 25
            ws2.column_dimensions['B'].width = 18
            ws2.column_dimensions['C'].width = 18
            ws2.column_dimensions['D'].width = 18
            ws2.column_dimensions['E'].width = 12

            # Salvar em buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            nome_arquivo = f"Comissoes_{meses_nomes_com[mes_com-1]}_{ano_com}.xlsx"
            st.download_button(
                label=f"📥 Baixar Relatório Excel — {meses_nomes_com[mes_com-1]}/{ano_com}",
                data=buffer,
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        except ImportError:
            st.warning("Instale openpyxl para exportar: pip install openpyxl")
    else:
        st.info(f"Nenhum item aprovado em {meses_nomes_com[mes_com-1]}/{ano_com}. Ajuste os filtros acima.")


# ==========================================
# TAB: CONFIGURAÇÕES
# ==========================================
with tab_config:
    config = load_config()

    st.markdown("#### ⚙️ Configurações do Sistema")

    st.markdown('<div class="section-title">🎯 Meta de Vendas</div>', unsafe_allow_html=True)
    nova_meta = st.number_input(
        "Meta mensal de vendas (R$)",
        min_value=0.0,
        value=float(config.get("meta_mensal", 50000)),
        step=1000.0,
        format="%.2f"
    )

    st.markdown('<div class="section-title">👥 Vendedores</div>', unsafe_allow_html=True)
    st.caption("Um vendedor por linha")
    vendedores_text = st.text_area(
        "Lista de vendedores",
        value="\n".join(config.get("vendedores", ["Allan"])),
        height=120
    )

    # 📸 Upload de fotos dos vendedores
    st.markdown('<div class="section-title">📸 Fotos dos Vendedores</div>', unsafe_allow_html=True)
    st.caption("Faça upload da foto de cada vendedor (aparece no ranking)")
    vendedores_atuais = [v.strip() for v in vendedores_text.split("\n") if v.strip()]
    fotos_atuais = config.get("vendedores_fotos", {})

    import base64
    fotos_novas = dict(fotos_atuais)
    for vend in vendedores_atuais:
        col_foto_preview, col_foto_upload = st.columns([1, 3])
        with col_foto_preview:
            avatar_html = get_avatar_html(vend, fotos_atuais)
            st.markdown(f'<div style="display:flex;align-items:center;height:60px;justify-content:center;">{avatar_html}</div>', unsafe_allow_html=True)
        with col_foto_upload:
            foto_file = st.file_uploader(
                f"Foto de {vend}",
                type=["png", "jpg", "jpeg"],
                key=f"foto_{vend}",
                label_visibility="collapsed"
            )
            if foto_file is not None:
                foto_bytes = foto_file.read()
                foto_b64 = base64.b64encode(foto_bytes).decode()
                ext = foto_file.type.split("/")[-1]
                fotos_novas[vend] = f"data:image/{ext};base64,{foto_b64}"

    if st.button("💾 Salvar Configurações", type="primary", use_container_width=True):
        vendedores_lista = [v.strip() for v in vendedores_text.split("\n") if v.strip()]
        config["meta_mensal"] = nova_meta
        config["vendedores"] = vendedores_lista
        config["vendedores_fotos"] = fotos_novas
        save_config(config)
        st.success("✅ Configurações salvas com sucesso!")
        st.rerun()
